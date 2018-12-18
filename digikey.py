from time import sleep
from openpyxl import load_workbook
from selenium import webdriver

orderFile = "digikeyitems.xlsx"
browser = webdriver.Chrome()


def excel2dict(file):
    digikeyOrder = {}
    wb = load_workbook(file)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        digikeyOrder[(sheet['C' + str(row)].value)] = sheet['D' + str(row)].value

    return digikeyOrder

def addItemsToCart(orderDict):
    for url,quantity in orderDict.items():
        if quantity == 0:
            continue

        browser.get(url)
        browser.find_element_by_name("qty").send_keys(quantity)
        browser.find_element_by_id("addtoorderbutton").click()

        sleep(2)

        try:
            browser.find_element_by_id("add-button").click()
        except:
            continue


orderDict = excel2dict(orderFile)
addItemsToCart(orderDict)




